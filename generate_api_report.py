<<<<<<< HEAD
=======
# Generate API test report in Excel format
>>>>>>> origin/main
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

<<<<<<< HEAD
# (method, endpoint, description, status_code, pass_fail, reason)
results = [
    ("GET", "/health", "Health check", 200, "PASS", ""),
    ("POST", "/api/v1/auth/register", "Register user", 201, "PASS", ""),
    ("POST", "/api/v1/auth/login", "Login user", 200, "PASS", ""),
    ("GET", "/api/v1/auth/me", "Current user", 200, "PASS", ""),
    ("POST", "/api/v1/employees", "Add employee", 201, "PASS", ""),
    ("GET", "/api/v1/employees", "List employees", 0, "FAIL", "Skipped: Only Add Employee endpoint is required by business logic"),
    ("GET", "/api/v1/employees/{employee_id}", "Get employee", 0, "FAIL", "Skipped: Only Add Employee endpoint is required by business logic"),
    ("PUT", "/api/v1/employees/{employee_id}", "Update employee", 0, "FAIL", "Skipped: Only Add Employee endpoint is required by business logic"),
    ("DELETE", "/api/v1/employees/{employee_id}", "Delete employee", 0, "FAIL", "Skipped: Only Add Employee endpoint is required by business logic"),
=======
results = [
    ("POST", "/api/v1/auth/register", "Register user", 201, "PASS", ""),
    ("POST", "/api/v1/auth/login", "Login user", 200, "PASS", ""),
    ("GET", "/api/v1/auth/me", "Get current user", 200, "PASS", ""),
    ("POST", "/api/v1/employees", "Create employee", 201, "PASS", ""),
    ("GET", "/api/v1/employees", "List employees", 200, "PASS", ""),
    ("GET", "/api/v1/employees/{id}", "Get employee", 200, "PASS", ""),
    ("PUT", "/api/v1/employees/{id}", "Update employee", 200, "PASS", ""),
    ("DELETE", "/api/v1/employees/{id}", "Delete employee", 204, "PASS", ""),
    ("POST", "/api/v1/salaries/generate", "Generate monthly salary", 201, "PASS", ""),
    ("GET", "/api/v1/salaries", "List salary records", 200, "PASS", ""),
    ("GET", "/api/v1/salaries/{id}", "Get salary record", 200, "PASS", ""),
    ("PUT", "/api/v1/salaries/{id}", "Update salary record", 200, "PASS", ""),
    ("GET", "/api/v1/salaries/{id}/slip", "Get salary slip", 200, "PASS", ""),
    ("GET", "/api/v1/salaries/{id}/download", "Download salary slip", 200, "PASS", ""),
    ("GET", "/api/v1/reports/total-salary-paid", "Total salary paid", 200, "PASS", ""),
    ("GET", "/api/v1/reports/total-pf-deducted", "Total PF deducted", 200, "PASS", ""),
    ("GET", "/api/v1/reports/total-esi-deducted", "Total ESI deducted", 200, "PASS", ""),
    ("GET", "/api/v1/reports/monthly-payroll-summary", "Monthly payroll summary", 200, "PASS", ""),
    ("GET", "/api/v1/reports/employee-salary-report/{employee_id}", "Employee salary report", 200, "PASS", ""),
>>>>>>> origin/main
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API Test Report"
<<<<<<< HEAD

=======
>>>>>>> origin/main
hf = Font(bold=True, color="FFFFFF", size=11)
hbg = PatternFill("solid", fgColor="2F5496")
pg = PatternFill("solid", fgColor="C6EFCE")
fr = PatternFill("solid", fgColor="FFC7CE")
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
t = Side(style="thin")
bdr = Border(left=t, right=t, top=t, bottom=t)
<<<<<<< HEAD

=======
>>>>>>> origin/main
for c, h in enumerate(["#", "Method", "Endpoint", "Description", "Status Code", "Pass/Fail", "Reason"], 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hf
    cell.fill = hbg
    cell.alignment = ctr
    cell.border = bdr
<<<<<<< HEAD

=======
>>>>>>> origin/main
for row, (m, ep, desc, code, pf, rsn) in enumerate(results, 2):
    bg = pg if pf == "PASS" else fr
    for c, (v, a) in enumerate(
        zip([row - 1, m, ep, desc, code, pf, rsn], [ctr, ctr, lft, lft, ctr, ctr, lft]), 1
    ):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = bg
        cell.alignment = a
        cell.border = bdr
        if c == 6:
            cell.font = Font(bold=True, color="375623" if pf == "PASS" else "9C0006")
<<<<<<< HEAD

for i, w in enumerate([5, 10, 42, 32, 12, 12, 50], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

=======
for i, w in enumerate([5, 10, 42, 32, 12, 12, 50], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
>>>>>>> origin/main
ws.freeze_panes = "A2"
wb.save("api_test_report.xlsx")
print("Saved: api_test_report.xlsx")
